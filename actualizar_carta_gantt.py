#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para actualizar la Carta Gantt con el estado de implementación hasta el 11-11-2025
"""

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from copy import copy

def actualizar_carta_gantt():
    archivo_original = "Carta_Gantt_Henko_Mochima_12_Oct_2025.xlsx"
    archivo_nuevo = f"Carta_Gantt_Henko_Mochima_Actualizada_{datetime.now().strftime('%d_%m_%Y_%H%M%S')}.xlsx"
    
    print("=" * 80)
    print("ACTUALIZANDO CARTA GANTT CON ESTADO DE IMPLEMENTACION")
    print("=" * 80)
    print()
    
    try:
        # Leer el archivo Excel original
        df = pd.read_excel(archivo_original, sheet_name="Gantt Henko & Mochima")
        
        # Identificar columnas
        col_tarea = [c for c in df.columns if 'TAREA' in str(c).upper()][0]
        col_responsable = [c for c in df.columns if 'RESPONSABLE' in str(c).upper()][0]
        col_estado = [c for c in df.columns if 'ESTADO' in str(c).upper()][0]
        col_inicio = [c for c in df.columns if 'INICIO' in str(c).upper()][0]
        col_fin = [c for c in df.columns if 'FIN' in str(c).upper()][0]
        
        # Definir las tareas implementadas y sus nuevos estados (SOLO hasta 11-11-2025)
        tareas_implementadas = {
            # Sprint 1
            "2.1 Sistema de login y seguridad": "Completada",
            "2.2 Dashboard principal Henko TCG y Mochima": "Completada",
            "2.3 Integración Excel": "Completada",
            "2.4 Eliminación del doble registro": "Completada",
            "2.5 Dashboard centro de control": "Completada",
            
            # Sprint 2
            "3.1 CRUD productos": "Completada",
            "3.2 Interfaz gestión inventarios": "Completada",
            "3.3 Sistema alertas stock mínimo": "Completada",
            "3.4 Consideración 7 días anticipación": "Completada",
            
            # Sprint 3 (SOLO hasta 11-11-2025)
            "4.1 Motor cálculo automático precios": "Completada",
        }
        
        # Tareas que NO se deben actualizar (están después del 11-11-2025 o no se implementaron)
        tareas_no_actualizar = [
            "2.6 Despliegue Sprint 1",
            "3.5 Despliegue Sprint 2",
            "4.2 Integración costos variables",
            "4.3 Recálculo márgenes dinámicos",
            "4.4 Interfaz automatización precios",
            "4.5 Despliegue Sprint 3",
            "5. SPRINT 4",
            "6. SPRINT 5",
        ]
        
        # Actualizar estados en el DataFrame
        actualizaciones = 0
        for idx, row in df.iterrows():
            tarea = str(row[col_tarea]) if pd.notna(row[col_tarea]) else ""
            
            # Limpiar caracteres especiales para comparación
            tarea_limpia = tarea.replace('→', '->').replace('–', '-').replace('—', '-')
            
            # Verificar que no esté en la lista de no actualizar
            no_actualizar = any(no_act in tarea_limpia for no_act in tareas_no_actualizar)
            if no_actualizar:
                continue
            
            # Buscar coincidencias parciales o exactas
            for tarea_key, nuevo_estado in tareas_implementadas.items():
                tarea_key_limpia = tarea_key.replace('→', '->').replace('–', '-').replace('—', '-')
                
                # Buscar coincidencia más específica
                palabras_clave = [p for p in tarea_key_limpia.lower().split() if len(p) > 3]
                coincidencias = sum(1 for palabra in palabras_clave if palabra in tarea_limpia.lower())
                
                # Requerir al menos 2 palabras clave para evitar falsos positivos
                if coincidencias >= 2 or tarea_key_limpia.lower() in tarea_limpia.lower():
                    if pd.notna(row[col_estado]):
                        estado_anterior = str(row[col_estado])
                        df.at[idx, col_estado] = nuevo_estado
                        actualizaciones += 1
                        try:
                            print(f"[ACTUALIZADO] {tarea_limpia[:60]}...")
                            print(f"   Estado anterior: {estado_anterior}")
                            print(f"   Estado nuevo: {nuevo_estado}")
                        except:
                            print(f"[ACTUALIZADO] Tarea actualizada")
                        break
        
        print()
        print(f"Total de tareas actualizadas: {actualizaciones}")
        print()
        
        # Actualizar el estado del Sprint 1
        for idx, row in df.iterrows():
            tarea = str(row[col_tarea]) if pd.notna(row[col_tarea]) else ""
            if "SPRINT 1 - AUTENTICACIÓN Y EXCEL" in tarea.upper():
                df.at[idx, col_estado] = "Completada"
                print(f"[ACTUALIZADO] Sprint 1: Completada")
            elif "SPRINT 2 - GESTIÓN INVENTARIOS" in tarea.upper():
                df.at[idx, col_estado] = "Completada"
                print(f"[ACTUALIZADO] Sprint 2: Completada")
            elif "SPRINT 3 - AUTOMATIZACIÓN PRECIOS" in tarea.upper():
                df.at[idx, col_estado] = "En progreso (20%)"  # Solo se completó 4.1 de 5 tareas (20%)
                print(f"[ACTUALIZADO] Sprint 3: En progreso (20%)")
        
        print()
        print("Guardando archivo actualizado...")
        
        # Guardar usando openpyxl para preservar formato
        wb = load_workbook(archivo_original)
        ws = wb["Gantt Henko & Mochima"]
        
        # Encontrar las columnas importantes
        estado_col = None
        tarea_col = None
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value and 'ESTADO' in str(cell.value).upper():
                estado_col = col_idx
            if cell.value and 'TAREA' in str(cell.value).upper():
                tarea_col = col_idx
        
        # Definir colores para cada sprint (título y subtareas)
        colores_sprints = {
            "1. ANÁLISIS Y DISEÑO": {
                "titulo": "0090EE90",  # Verde más intenso
                "subtareas": "00D5E8D6"  # Verde claro
            },
            "2. SPRINT 1": {
                "titulo": "0047B8D8",  # Azul más intenso
                "subtareas": "00B4D8F7"  # Azul claro
            },
            "3. SPRINT 2": {
                "titulo": "00E8A87C",  # Naranja más intenso
                "subtareas": "00E8D5B7"  # Naranja claro
            },
            "4. SPRINT 3": {
                "titulo": "00F7B87C",  # Melocotón más intenso
                "subtareas": "00F7D5B4"  # Melocotón claro
            },
            "5. SPRINT 4": {
                "titulo": "00D5A4E8",  # Morado más intenso
                "subtareas": "00D5B4E8"  # Morado claro
            },
            "6. SPRINT 5": {
                "titulo": "00E8A4D5",  # Rosa más intenso
                "subtareas": "00E8B4D5"  # Rosa claro
            },
        }
        
        if estado_col:
            # Actualizar estados en el archivo Excel
            for idx, row in df.iterrows():
                tarea = str(row[col_tarea]) if pd.notna(row[col_tarea]) else ""
                nuevo_estado = row[col_estado] if pd.notna(row[col_estado]) else None
                
                if nuevo_estado and tarea:
                    # Buscar la fila en el Excel (índice + 2 porque Excel empieza en 1 y tiene header)
                    excel_row = idx + 2
                    cell = ws.cell(row=excel_row, column=estado_col)
                    
                    # Actualizar el valor
                    if cell.value != nuevo_estado:
                        cell.value = nuevo_estado
                        
                        # Aplicar formato según estado
                        if "Completada" in str(nuevo_estado):
                            cell.fill = PatternFill(start_color="00C6EFCE", end_color="00C6EFCE", fill_type="solid")
                            cell.font = Font(bold=True, color="00000000")
                        elif "En progreso" in str(nuevo_estado):
                            cell.fill = PatternFill(start_color="00FFC000", end_color="00FFC000", fill_type="solid")
                            cell.font = Font(bold=True, color="00000000")
                        elif "Pendiente" in str(nuevo_estado):
                            cell.fill = PatternFill(start_color="00D9D9D9", end_color="00D9D9D9", fill_type="solid")
                            cell.font = Font(bold=True, color="00000000")
        
        # Colorear filas de sprints
        if tarea_col:
            print("Aplicando colores a los sprints...")
            sprint_actual = None
            sprint_row_start = None
            
            for row_idx in range(2, ws.max_row + 1):
                tarea_cell = ws.cell(row=row_idx, column=tarea_col)
                tarea_value = str(tarea_cell.value) if tarea_cell.value else ""
                
                # Identificar si es un título de sprint
                es_titulo_sprint = False
                for sprint_key, colores in colores_sprints.items():
                    if sprint_key.upper() in tarea_value.upper() and not tarea_value.strip().startswith(('1.1', '1.2', '1.3', '2.1', '2.2', '2.3', '2.4', '2.5', '2.6', '3.1', '3.2', '3.3', '3.4', '3.5', '4.1', '4.2', '4.3', '4.4', '4.5', '5.1', '5.2', '5.3', '5.4', '5.5', '6.1', '6.2')):
                        sprint_actual = sprint_key
                        sprint_row_start = row_idx
                        es_titulo_sprint = True
                        
                        # Colorear título del sprint con color más intenso
                        for col_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            if col_idx != estado_col or not cell.fill.start_color:
                                cell.fill = PatternFill(start_color=colores["titulo"], end_color=colores["titulo"], fill_type="solid")
                                cell.font = Font(bold=True, color="00FFFFFF")  # Texto blanco para mejor contraste
                        print(f"  [COLOR] Título Sprint: {sprint_key} - Fila {row_idx} (color intenso)")
                        break
                
                # Si no es título pero hay un sprint activo, colorear como subtarea
                if not es_titulo_sprint and sprint_actual and sprint_row_start:
                    # Verificar que sea una subtarea del sprint actual
                    numero_tarea = tarea_value.split('.')[0] if '.' in tarea_value else ""
                    if numero_tarea and sprint_actual.split('.')[0] in numero_tarea:
                        colores = colores_sprints[sprint_actual]
                        # Colorear subtarea con color más claro
                        for col_idx in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            # No sobrescribir el color de estado
                            if col_idx != estado_col:
                                cell.fill = PatternFill(start_color=colores["subtareas"], end_color=colores["subtareas"], fill_type="solid")
                                if not cell.font or not cell.font.bold:
                                    cell.font = Font(bold=False, color="00000000")
                    else:
                        # Si no coincide, resetear el sprint actual
                        sprint_actual = None
                        sprint_row_start = None
        
        # Guardar el archivo nuevo
        wb.save(archivo_nuevo)
        print(f"[OK] Archivo guardado: {archivo_nuevo}")
        print()
        print("=" * 80)
        print("RESUMEN DE ACTUALIZACIONES")
        print("=" * 80)
        print()
        print("TAREAS ACTUALIZADAS A 'COMPLETADA':")
        print("  - 2.1 Sistema de login y seguridad")
        print("  - 2.2 Dashboard principal Henko TCG y Mochima")
        print("  - 2.3 Integracion Excel -> Base de datos cloud")
        print("  - 2.4 Eliminacion del doble registro")
        print("  - 2.5 Dashboard centro de control para KPIs")
        print("  - 3.1 CRUD productos")
        print("  - 3.2 Interfaz gestion inventarios")
        print("  - 3.3 Sistema alertas stock minimo (HU006)")
        print("  - 3.4 Consideracion 7 dias anticipacion Henko")
        print("  - 4.1 Motor calculo automatico precios (HU002)")
        print()
        print("SPRINTS ACTUALIZADOS:")
        print("  - Sprint 1: Completada")
        print("  - Sprint 2: Completada")
        print("  - Sprint 3: En progreso (20%) - Solo tarea 4.1 completada")
        print()
        
    except Exception as e:
        print(f"[ERROR] {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    actualizar_carta_gantt()
