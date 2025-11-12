#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para actualizar la Carta Gantt con barras de avance generadas automáticamente
Versión: 11-11-2025
"""

import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from copy import copy

def parse_fecha(fecha_valor):
    """Convierte un valor de fecha a datetime"""
    if pd.isna(fecha_valor):
        return None
    if isinstance(fecha_valor, datetime):
        return fecha_valor
    if isinstance(fecha_valor, str):
        # Intentar varios formatos
        formatos = ['%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y']
        for formato in formatos:
            try:
                return datetime.strptime(fecha_valor, formato)
            except:
                continue
    return None

def generar_columnas_fechas(fecha_inicio_proyecto, fecha_fin_proyecto, intervalo='semanal'):
    """
    Genera una lista de fechas para las columnas del Gantt
    intervalo puede ser: 'diario', 'semanal', 'quincenal', 'mensual'
    """
    fechas = []
    fecha_actual = fecha_inicio_proyecto
    
    if intervalo == 'semanal':
        delta = timedelta(days=7)
    elif intervalo == 'quincenal':
        delta = timedelta(days=14)
    elif intervalo == 'mensual':
        delta = timedelta(days=30)
    else:  # diario
        delta = timedelta(days=1)
    
    while fecha_actual <= fecha_fin_proyecto:
        fechas.append(fecha_actual)
        fecha_actual += delta
    
    return fechas

def actualizar_carta_gantt():
    import glob
    import os
    
    # Buscar el archivo más reciente de la Carta Gantt
    archivos_gantt = glob.glob("Carta_Gantt*.xlsx")
    if not archivos_gantt:
        print("[ERROR] No se encontró ningún archivo de Carta Gantt")
        return
    
    # Usar el más reciente
    archivo_original = max(archivos_gantt, key=os.path.getmtime)
    archivo_nuevo = f"Carta_Gantt_Henko_Mochima_Con_Barras_{datetime.now().strftime('%d_%m_%Y_%H%M%S')}.xlsx"
    
    print(f"Usando archivo base: {archivo_original}")
    
    print("=" * 80)
    print("ACTUALIZANDO CARTA GANTT CON BARRAS DE AVANCE")
    print("=" * 80)
    print()
    
    try:
        # Leer el archivo Excel original
        df = pd.read_excel(archivo_original, sheet_name="Gantt Henko & Mochima")
        
        # Identificar columnas
        col_tarea = [c for c in df.columns if 'TAREA' in str(c).upper()][0]
        col_responsable = [c for c in df.columns if 'RESPONSABLE' in str(c).upper()][0]
        col_estado = [c for c in df.columns if 'ESTADO' in str(c).upper()][0]
        col_inicio = [c for c in df.columns if 'INICIO' in str(c).upper()][0]
        col_fin = [c for c in df.columns if 'FIN' in str(c).upper()][0]
        
        print(f"✅ Columnas identificadas:")
        print(f"   - TAREA: {col_tarea}")
        print(f"   - ESTADO: {col_estado}")
        print(f"   - INICIO: {col_inicio}")
        print(f"   - FIN: {col_fin}")
        print()
        
        # Definir las tareas implementadas y sus estados
        tareas_implementadas_hoy = {
            "4.4 Interfaz automatización precios": "Completada",
            "4.4 Configuración personalizable márgenes": "Completada",
            "4.4 Configuración personalizable categorías": "Completada",
            "4.4 Aplicación precios recomendados por unidad": "Completada",
            "4.4 Visualización categoría y margen en tabla": "Completada",
            "4.4 Optimización cálculo salud inventario": "Completada",
            "4.4 Botón mejorar salud inventario": "Completada",
        }
        
        tareas_ya_existentes = {
            "2.1 Sistema de login y seguridad": "Completada",
            "2.2 Dashboard principal Henko TCG y Mochima": "Completada",
            "2.3 Integración Excel": "Completada",
            "2.4 Eliminación del doble registro": "Completada",
            "2.5 Dashboard centro de control": "Completada",
            "3.1 CRUD productos": "Completada",
            "3.2 Interfaz gestión inventarios": "Completada",
            "3.3 Sistema alertas stock mínimo": "Completada",
            "3.4 Consideración 7 días anticipación": "Completada",
            "4.1 Motor cálculo automático precios": "Completada",
        }
        
        tareas_implementadas = {**tareas_implementadas_hoy, **tareas_ya_existentes}
        
        # Actualizar estados en el DataFrame
        actualizaciones = 0
        for idx, row in df.iterrows():
            tarea = str(row[col_tarea]) if pd.notna(row[col_tarea]) else ""
            tarea_limpia = tarea.replace('→', '->').replace('–', '-').replace('—', '-')
            
            for tarea_key, nuevo_estado in tareas_implementadas.items():
                tarea_key_limpia = tarea_key.replace('→', '->').replace('–', '-').replace('—', '-')
                palabras_clave = [p for p in tarea_key_limpia.lower().split() if len(p) > 3]
                coincidencias = sum(1 for palabra in palabras_clave if palabra in tarea_limpia.lower())
                
                if coincidencias >= 2 or tarea_key_limpia.lower() in tarea_limpia.lower():
                    if pd.notna(row[col_estado]):
                        estado_anterior = str(row[col_estado])
                        if estado_anterior != nuevo_estado:
                            df.at[idx, col_estado] = nuevo_estado
                            actualizaciones += 1
                            print(f"[ACTUALIZADO] {tarea_limpia[:60]}... -> {nuevo_estado}")
                        break
        
        print(f"\nTotal de tareas actualizadas: {actualizaciones}\n")
        
        # Actualizar estados de sprints
        for idx, row in df.iterrows():
            tarea = str(row[col_tarea]) if pd.notna(row[col_tarea]) else ""
            if "SPRINT 1" in tarea.upper() and "AUTENTICACIÓN" in tarea.upper():
                df.at[idx, col_estado] = "Completada"
            elif "SPRINT 2" in tarea.upper() and "INVENTARIOS" in tarea.upper():
                df.at[idx, col_estado] = "Completada"
            elif "SPRINT 3" in tarea.upper() and "AUTOMATIZACIÓN" in tarea.upper():
                df.at[idx, col_estado] = "En progreso (90%)"
        
        # Abrir con openpyxl para trabajar con formato
        wb = load_workbook(archivo_original)
        ws = wb["Gantt Henko & Mochima"]
        
        # Encontrar columnas en Excel
        estado_col = None
        tarea_col = None
        inicio_col = None
        fin_col = None
        
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                valor = str(cell.value).upper()
                if 'ESTADO' in valor:
                    estado_col = col_idx
                if 'TAREA' in valor:
                    tarea_col = col_idx
                if 'INICIO' in valor:
                    inicio_col = col_idx
                if 'FIN' in valor:
                    fin_col = col_idx
        
        primera_col_gantt = fin_col + 1
        
        print(f"🔍 Columnas detectadas en Excel:")
        print(f"   TAREA: {tarea_col}")
        print(f"   ESTADO: {estado_col}")
        print(f"   INICIO: {inicio_col}")
        print(f"   FIN: {fin_col}")
        print(f"   Primera columna Gantt: {primera_col_gantt}")
        print()
        
        # Leer las fechas del header para las columnas del Gantt
        fechas_columnas = {}
        for col_idx in range(primera_col_gantt, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col_idx)
            if header_cell.value:
                # Las fechas están en el header, pueden ser datetime o strings con "###"
                if isinstance(header_cell.value, datetime):
                    fechas_columnas[col_idx] = header_cell.value
                elif isinstance(header_cell.value, str) and '###' in header_cell.value:
                    # Son marcadores de semana/período
                    fechas_columnas[col_idx] = col_idx  # Usar índice como referencia
        
        print(f"📅 Se detectaron {len(fechas_columnas)} columnas de fecha en el Gantt")
        print()
        
        # Colores para los diferentes estados
        colores_estado = {
            "Completada": "0000B050",  # Verde
            "En progreso": "00FFC000",  # Naranja
            "Pendiente": "004472C4",   # Azul
        }
        
        # Colores para sprints (fondos de filas)
        colores_sprints = {
            "1. ANÁLISIS Y DISEÑO": {"titulo": "0090EE90", "subtareas": "00D5E8D6"},
            "2. SPRINT 1": {"titulo": "0047B8D8", "subtareas": "00B4D8F7"},
            "3. SPRINT 2": {"titulo": "00E8A87C", "subtareas": "00E8D5B7"},
            "4. SPRINT 3": {"titulo": "00F7B87C", "subtareas": "00F7D5B4"},
            "5. SPRINT 4": {"titulo": "00D5A4E8", "subtareas": "00D5B4E8"},
            "6. SPRINT 5": {"titulo": "00E8A4D5", "subtareas": "00E8B4D5"},
        }
        
        print("🎨 Generando barras de avance del Gantt...")
        print()
        
        # Generar las barras de avance para cada tarea
        barras_generadas = 0
        for row_idx in range(2, ws.max_row + 1):
            tarea_cell = ws.cell(row=row_idx, column=tarea_col)
            inicio_cell = ws.cell(row=row_idx, column=inicio_col)
            fin_cell = ws.cell(row=row_idx, column=fin_col)
            estado_cell = ws.cell(row=row_idx, column=estado_col)
            
            tarea = str(tarea_cell.value) if tarea_cell.value else ""
            inicio = inicio_cell.value
            fin = fin_cell.value
            estado = str(estado_cell.value) if estado_cell.value else "Pendiente"
            
            # Saltar filas de títulos de sprint (no tienen fechas)
            if not inicio or not fin:
                continue
            
            # Convertir fechas
            fecha_inicio = parse_fecha(inicio)
            fecha_fin = parse_fecha(fin)
            
            if not fecha_inicio or not fecha_fin:
                continue
            
            # Determinar el color según el estado
            color_barra = colores_estado.get("Completada", "0000B050")
            if "progreso" in estado.lower():
                color_barra = colores_estado.get("En progreso", "00FFC000")
            elif "pendiente" in estado.lower():
                color_barra = colores_estado.get("Pendiente", "004472C4")
            
            # Pintar las celdas correspondientes al período de la tarea
            # Como no tenemos las fechas exactas del header, usamos una aproximación
            # basada en la posición relativa de las fechas
            
            # Obtener todas las fechas del proyecto
            todas_fechas_inicio = []
            todas_fechas_fin = []
            for r in range(2, ws.max_row + 1):
                ini = ws.cell(row=r, column=inicio_col).value
                fi = ws.cell(row=r, column=fin_col).value
                if ini:
                    f_ini = parse_fecha(ini)
                    if f_ini:
                        todas_fechas_inicio.append(f_ini)
                if fi:
                    f_fi = parse_fecha(fi)
                    if f_fi:
                        todas_fechas_fin.append(f_fi)
            
            if todas_fechas_inicio and todas_fechas_fin:
                fecha_min_proyecto = min(todas_fechas_inicio)
                fecha_max_proyecto = max(todas_fechas_fin)
                duracion_proyecto = (fecha_max_proyecto - fecha_min_proyecto).days
                
                # Calcular posición relativa de la tarea
                inicio_relativo = (fecha_inicio - fecha_min_proyecto).days / duracion_proyecto
                fin_relativo = (fecha_fin - fecha_min_proyecto).days / duracion_proyecto
                
                # Número de columnas disponibles para el Gantt
                num_cols_gantt = ws.max_column - primera_col_gantt + 1
                
                # Calcular las columnas a pintar
                col_inicio_barra = primera_col_gantt + int(inicio_relativo * num_cols_gantt)
                col_fin_barra = primera_col_gantt + int(fin_relativo * num_cols_gantt)
                
                # Pintar las celdas
                for col_idx in range(col_inicio_barra, min(col_fin_barra + 1, ws.max_column + 1)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill(start_color=color_barra, end_color=color_barra, fill_type="solid")
                
                barras_generadas += 1
                print(f"  ✅ Barra generada: {tarea[:50]}... ({fecha_inicio.strftime('%d-%m')} a {fecha_fin.strftime('%d-%m')})")
        
        print(f"\n📊 Total de barras generadas: {barras_generadas}\n")
        
        # Actualizar estados en el Excel
        if estado_col:
            print("🔄 Actualizando estados...")
            for idx, row in df.iterrows():
                nuevo_estado = row[col_estado] if pd.notna(row[col_estado]) else None
                if nuevo_estado:
                    excel_row = idx + 2
                    cell = ws.cell(row=excel_row, column=estado_col)
                    cell.value = nuevo_estado
                    
                    # Aplicar formato según estado
                    if "Completada" in str(nuevo_estado):
                        cell.fill = PatternFill(start_color="00C6EFCE", end_color="00C6EFCE", fill_type="solid")
                        cell.font = Font(bold=True, color="00FFFFFF")
                    elif "progreso" in str(nuevo_estado).lower():
                        cell.fill = PatternFill(start_color="00FFC000", end_color="00FFC000", fill_type="solid")
                        cell.font = Font(bold=True, color="00FFFFFF")
                    elif "Pendiente" in str(nuevo_estado):
                        cell.fill = PatternFill(start_color="00D9D9D9", end_color="00D9D9D9", fill_type="solid")
                        cell.font = Font(bold=True, color="00FFFFFF")
        
        # Colorear filas de sprints (SOLO en columnas de datos, no en Gantt)
        if tarea_col:
            print("🎨 Aplicando colores a sprints...")
            sprint_actual = None
            
            for row_idx in range(2, ws.max_row + 1):
                tarea_cell = ws.cell(row=row_idx, column=tarea_col)
                tarea_value = str(tarea_cell.value) if tarea_cell.value else ""
                tarea_value_clean = tarea_value.strip()
                
                # Identificar títulos de sprint
                es_titulo_sprint = False
                for sprint_key, colores in colores_sprints.items():
                    if sprint_key.upper() in tarea_value_clean.upper():
                        if not any(tarea_value_clean.startswith(f"{i}.{j}") for i in range(1, 7) for j in range(1, 10)):
                            sprint_actual = sprint_key
                            es_titulo_sprint = True
                            
                            # Colorear título SOLO en columnas de datos
                            for col_idx in range(1, primera_col_gantt):
                                cell = ws.cell(row=row_idx, column=col_idx)
                                if col_idx != estado_col:
                                    cell.fill = PatternFill(start_color=colores["titulo"], end_color=colores["titulo"], fill_type="solid")
                                    cell.font = Font(bold=True, color="00FFFFFF")
                            break
                
                # Colorear subtareas SOLO en columnas de datos
                if not es_titulo_sprint and sprint_actual:
                    numero_tarea = tarea_value_clean.split('.')[0] if '.' in tarea_value_clean else ""
                    if numero_tarea and sprint_actual.split('.')[0] in numero_tarea:
                        colores = colores_sprints[sprint_actual]
                        for col_idx in range(1, primera_col_gantt):
                            if col_idx != estado_col:
                                cell = ws.cell(row=row_idx, column=col_idx)
                                cell.fill = PatternFill(start_color=colores["subtareas"], end_color=colores["subtareas"], fill_type="solid")
                                cell.font = Font(color="00FFFFFF")
        
        # Guardar archivo
        wb.save(archivo_nuevo)
        print(f"\n✅ Archivo guardado: {archivo_nuevo}")
        print()
        print("=" * 80)
        print("RESUMEN")
        print("=" * 80)
        print(f"✅ {actualizaciones} estados actualizados")
        print(f"✅ {barras_generadas} barras de Gantt generadas")
        print(f"✅ Sprints coloreados")
        print(f"✅ Archivo: {archivo_nuevo}")
        print()
        
    except Exception as e:
        print(f"[ERROR] {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    actualizar_carta_gantt()